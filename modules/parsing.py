"""
modules/parsing.py
Excel / CSV ingestion: classify sheet type, parse rows into list-of-dicts,
skip aggregate/totals rows.

Handles both clean structured layouts and messy legacy "print-style" formats
where:
  - Column headers span two rows (group label + sub-label)
  - Each claim has a sub-row with the address (col B) and cause-of-loss (col D)
  - Section separator rows contain only '----------' dashes
  - Section subtotal rows start with 'Total ...'

NEW (title extraction):
  extract_sheet_title_kvs() scans the pre-header rows of any sheet and returns
  a dict of canonical key → value pairs (TPA Name, Treaty, Cedant,
  Valuation Date, Sheet Title, Sheet Name …).  extract_from_excel() now
  returns this dict as a third element so callers can pass it downstream to
  schema_mapping.extract_title_fields_from_kvs().

CHANGED (smart sub-row inference):
  Sub-rows in legacy print layouts are no longer parsed by hardcoded column
  positions (col B = Address, col D = Cause of Loss).  Instead every non-empty
  cell on a sub-row is classified by value pattern:
    • Street address pattern      → "Address"
    • City / State / ZIP pattern  → "City State Zip"
    • Known peril / event words   → "Cause of Loss"
    • Proper-noun name pattern    → "Claimant Name"
    • Unknown + col B position    → "Address"        (legacy fallback)
    • Unknown + col D position    → "Cause of Loss"  (legacy fallback)
    • Unknown + any other col     → "SubRow_<col>"   (data preserved)
  This allows the parser to handle sheets where address/cause appear in
  non-standard column positions.
"""

import csv
import os
import re

import openpyxl

from modules.cell_format import format_cell_value_with_fmt


# ── Sheet classifier ──────────────────────────────────────────────────────────

def classify_sheet(rows) -> str:
    text = " ".join(str(cell).lower() for row in rows[:20] for cell in row if cell)

    # SUMMARY detection: "line of business" must co-occur with summary-specific
    # signals, not just appear as a data column header in a loss-run sheet.
    if "line of business" in text:
        summary_co_signals = [
            "# claims", "num claims", "number of claims", "claim count",
            "loss ratio", "loss rate", "frequency", "severity",
        ]
        if any(sig in text for sig in summary_co_signals):
            return "SUMMARY"
        # Also SUMMARY if "line of business" is a standalone first-cell row label
        for row in rows[:20]:
            non_empty = [v for v in row if v is not None and str(v).strip()]
            if non_empty and str(non_empty[0]).lower().strip() == "line of business" and len(non_empty) == 1:
                return "SUMMARY"

    has_claim = any(x in text for x in [
        "claim number", "claim no", "claim #", "claim id", "claim_id",
        "claim ref", "claimant", "file number", "file no", "file num",
        "file ref",
    ])
    has_loss = any(x in text for x in [
        "loss date", "date of loss", "loss dt", "accident date",
        "occurrence date", "incident date", "date of injury", "date of incident",
        "injury date", "dol",
    ])
    has_fin = any(x in text for x in [
        "incurred", "paid", "reserve", "outstanding",
        "total paid", "total incurred", "indemnity", "expense",
    ])
    if has_claim and (has_loss or has_fin):
        return "LOSS_RUN"
    if "policy" in text and ("claim" in text or "incurred" in text):
        return "COMMERCIAL_LOSS_RUN"
    if has_claim:
        return "LOSS_RUN"
    return "UNKNOWN"


# ── Legacy-layout detector ────────────────────────────────────────────────────

def _is_legacy_print_layout(rows: list) -> bool:
    """
    Detect a legacy print-style TPA loss run layout.
    Signatures:
      - Two adjacent rows that together form the column header (neither row
        alone passes _find_header_row's threshold, but together they do).
      - Rows consisting entirely of '----------' dashes appear in the data area.
      - Sub-rows where only columns B and/or D have values (address + cause).
    """
    # Check for '----------' separator rows
    for row in rows:
        non_empty = [c for c in row if c is not None]
        if non_empty and all(str(c).strip() == "----------" for c in non_empty):
            return True

    # Check for the characteristic 2-row header pattern:
    # row N has group labels in some cols, row N+1 has sub-labels in ALL cols
    for i in range(min(20, len(rows) - 1)):
        r1_vals = [str(c).strip() for c in rows[i] if c is not None]
        r2_vals = [str(c).strip() for c in rows[i + 1] if c is not None]
        if len(r2_vals) >= 5 and len(r1_vals) >= 2:
            r1_filled = sum(1 for c in rows[i] if c)
            r2_filled = sum(1 for c in rows[i + 1] if c)
            if r2_filled > r1_filled * 1.5 and r1_filled >= 2:
                combined = " ".join(r1_vals + r2_vals).lower()
                if ("file" in combined or "claim" in combined) and (
                    "paid" in combined or "incurred" in combined or "outstanding" in combined
                ):
                    return True
    return False


def _find_legacy_header_rows(rows: list) -> tuple[int, int] | None:
    """
    For legacy layouts find the two consecutive rows that form the header.
    Returns (row_index_of_group_label_row, row_index_of_sub_label_row) or None.
    """
    for i in range(min(25, len(rows) - 1)):
        r1 = rows[i]
        r2 = rows[i + 1]
        r1_filled = sum(1 for c in r1 if c)
        r2_filled = sum(1 for c in r2 if c)
        if r2_filled < 4:
            continue
        combined = " ".join(
            str(c).lower() for c in list(r1) + list(r2) if c
        )
        if ("file" in combined or "claim" in combined or "assured" in combined) and (
            "paid" in combined or "outstanding" in combined or "incurred" in combined
        ):
            if r1_filled >= 2:
                return (i, i + 1)
            if r2_filled >= 5:
                return (i + 1, i + 1)
    return None


def _merge_two_header_rows(row1: list, row2: list) -> list[str]:
    """
    Combine a group-label row and a sub-label row into one list of column names.
    Duplicate merged names get a numeric suffix (_2, _3 …).
    """
    headers: list[str] = []
    seen: dict[str, int] = {}
    for g, s in zip(row1, row2):
        g_s = str(g).strip() if g else ""
        s_s = str(s).strip() if s else ""
        if g_s and s_s and g_s.upper() != s_s.upper():
            name = f"{g_s} {s_s}"
        elif s_s:
            name = s_s
        elif g_s:
            name = g_s
        else:
            name = ""
        if name:
            seen[name] = seen.get(name, 0) + 1
            if seen[name] > 1:
                name = f"{name}_{seen[name]}"
        headers.append(name)
    return headers


# ── Sub-row / separator / subtotal detectors ─────────────────────────────────

def _is_separator_row(row_values: list) -> bool:
    """Row containing only '----------' dashes (and empty cells)."""
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if not non_empty:
        return False
    return all(str(c).strip() == "----------" for c in non_empty)


def _is_subtotal_row(row_values: list) -> bool:
    """Row whose first non-empty cell starts with 'Total' (section subtotal)."""
    for c in row_values:
        if c is not None and str(c).strip():
            return bool(re.match(r"^total\b", str(c).strip(), re.IGNORECASE))
    return False


def _is_legacy_sub_row(row_values: list, num_cols: int) -> bool:
    """
    In legacy print layouts each claim is followed by a sub-row containing
    only the address (col B / index 1) and/or cause-of-loss (col D / index 3).
    Signature: col 0 (FILE NUM) is empty, ≤ 3 non-empty cells total, and at
    least one of col 1 or col 3 has a value.
    """
    if not row_values or row_values[0] is not None:
        return False
    non_empty = [c for c in row_values if c is not None and str(c).strip()]
    if len(non_empty) == 0 or len(non_empty) > 3:
        return False
    has_addr_or_cause = (
        (len(row_values) > 1 and row_values[1] is not None) or
        (len(row_values) > 3 and row_values[3] is not None)
    )
    return has_addr_or_cause


# ── Smart sub-row cell classifier ─────────────────────────────────────────────

# Street address: starts with a number followed by a street name
_ADDRESS_PAT = re.compile(
    r"""
    ^(
        \d+\s+\w.*            # "391 MAIN ST"
      | P\.?O\.?\s*BOX\s+\d   # "PO BOX 443"
      | \d+[-/]\d+\s+\w.*     # "12-14 ELM RD"
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)

# Street suffix keywords that strongly imply an address line
_ADDRESS_SUFFIX = re.compile(
    r"\b(st|street|ave|avenue|blvd|boulevard|rd|road|dr|drive|"
    r"ln|lane|ct|court|pl|place|way|cir|circle|hwy|highway|"
    r"pkwy|parkway|terr|ter|loop|trail|trl|run|box|suite|ste|"
    r"apt|unit|floor|fl)\b",
    re.IGNORECASE,
)

# City / State / ZIP  — "AUSTIN TX 78701" or "AUSTIN, TX" or "78701"
_CITY_STATE_ZIP_PAT = re.compile(
    r"""
    ^(
        [A-Za-z\s]{2,30},?\s+[A-Z]{2}\s+\d{5}(-\d{4})?   # City, ST 12345
      | [A-Za-z\s]{2,30},?\s+[A-Z]{2}$                     # City, ST
      | \d{5}(-\d{4})?$                                     # ZIP only
    )
    """,
    re.IGNORECASE | re.VERBOSE,
)

# Cause of loss: typical peril / event words
_CAUSE_OF_LOSS_PAT = re.compile(
    r"\b(fire|flood|wind|windstorm|hail|storm|tornado|hurricane|"
    r"tropical\s+storm|water\s+damage|water\s+intrusion|theft|"
    r"vandalism|slip|fall|trip|collision|accident|explosion|"
    r"lightning|freeze|ice|snow|earthquake|sinkhole|mold|"
    r"liability|negligence|assault|discrimination|wrongful|"
    r"product\s+liability|premises|auto|vehicle|medical|workers|"
    r"comp|injury|glass|burst\s+pipe|pipe\s+burst|roof|damage)\b",
    re.IGNORECASE,
)

# Name pattern: two or more capitalised words, no digits
_NAME_PAT = re.compile(
    r"^[A-Z][A-Za-z'-]+(\s+[A-Z][A-Za-z'-]+){1,4}$"
)


def _col_letter(col_index: int) -> str:
    """Convert 0-based column index to Excel letter (0→A, 1→B, …)."""
    result = ""
    n = col_index + 1
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def _classify_subrow_cell(value: str) -> str:
    """
    Classify a single sub-row cell value into a semantic field name.

    Returns one of:
      "Address"        – looks like a street address line
      "City State Zip" – looks like city/state/zip continuation
      "Cause of Loss"  – matches known peril / event vocabulary
      "Claimant Name"  – looks like a person/company name
      "Unknown"        – cannot be confidently classified
    """
    v = str(value).strip()
    if not v:
        return "Unknown"

    if _ADDRESS_PAT.match(v):
        return "Address"
    if _ADDRESS_SUFFIX.search(v) and re.search(r'\d', v):
        return "Address"
    if _CITY_STATE_ZIP_PAT.match(v):
        return "City State Zip"
    if _CAUSE_OF_LOSS_PAT.search(v):
        return "Cause of Loss"
    if _NAME_PAT.match(v) and not re.search(r'\d', v):
        return "Claimant Name"
    return "Unknown"


def _infer_subrow_fields(raw_row: list) -> dict[str, tuple[str, int]]:
    """
    Given a raw sub-row (list of cell values), return a dict mapping
    inferred field names → (value, 1-based col index).

    Strategy:
      1. Classify every non-empty cell by _classify_subrow_cell().
      2. If a field type was confidently identified, use it.
      3. If classification is "Unknown":
           - col index 1 (B) → fall back to "Address"        (legacy positional)
           - col index 3 (D) → fall back to "Cause of Loss"  (legacy positional)
           - other cols      → store as "SubRow_<col_letter>" so data is
                               not silently dropped
      4. If a field type is claimed by two cells, the second gets a _2 suffix.
    """
    result: dict[str, tuple[str, int]] = {}
    type_count: dict[str, int] = {}

    for c_idx, val in enumerate(raw_row):
        if val is None or not str(val).strip():
            continue

        val_s      = str(val).strip()
        field_type = _classify_subrow_cell(val_s)

        # Positional fallback for "Unknown"
        if field_type == "Unknown":
            if c_idx == 1:
                field_type = "Address"
            elif c_idx == 3:
                field_type = "Cause of Loss"
            else:
                field_type = f"SubRow_{_col_letter(c_idx)}"

        # Handle duplicate field types
        type_count[field_type] = type_count.get(field_type, 0) + 1
        if type_count[field_type] > 1:
            field_type = f"{field_type}_{type_count[field_type]}"

        result[field_type] = (val_s, c_idx + 1)  # 1-based col

    return result


def _enrich_from_subrow(
    claim: dict,
    raw_row: list,
    r_idx: int,
) -> None:
    """
    Smart replacement for the hardcoded col-B/col-D sub-row enrichment.

    Infers field names from sub-row cell values using pattern matching and
    calls _enrich_field() for each one.  Falls back gracefully to legacy
    positional logic for values that cannot be pattern-classified.
    """
    inferred = _infer_subrow_fields(raw_row)
    for field_name, (value, excel_col) in inferred.items():
        _enrich_field(claim, field_name, value, excel_row=r_idx, excel_col=excel_col)


# ── Aggregate-row detection ───────────────────────────────────────────────────

_AGGREGATE_PATTERNS = re.compile(
    r"^(total|totals|grand\s*total|subtotal|aggregate|summary|sum|report\s*(date|total|summary)|"
    r"all\s+adjusters|ytd\s+total|period\s+total|fiscal\s+total|portfolio\s+total|"
    r"TOTALS_AGGREGATE|SUMMARY_FLIBBER|AGGREGATE_ZORP|SUMMARY_ZORP)",
    re.IGNORECASE,
)
_AGGREGATE_EXTRA = re.compile(
    r"(aggregate|zorp|flibber|summary|zoop|gorp|totals?_|_total|report_date|all_adjuster)",
    re.IGNORECASE,
)


def _is_aggregate_row(row_values: list) -> bool:
    non_empty = [str(v).strip() for v in row_values if v is not None and str(v).strip()]
    if not non_empty:
        return False
    first_val = non_empty[0]
    if _AGGREGATE_PATTERNS.match(first_val):
        return True
    if _AGGREGATE_EXTRA.search(first_val):
        return True
    first_tokens     = re.split(r"[_\s]+", first_val.lower())
    aggregate_tokens = {"total", "totals", "aggregate", "summary", "subtotal", "grand", "portfolio", "report"}
    if len(first_tokens) >= 2 and any(t in aggregate_tokens for t in first_tokens):
        return True
    for v in non_empty[:6]:
        if re.match(
            r"(total\s+claims|report\s+date|all\s+adjusters|open:\s*\d|pend:\s*\d|open:\d)",
            str(v), re.IGNORECASE,
        ):
            return True
    nums = [float(v) for v in row_values if isinstance(v, (int, float))]
    if nums and len(nums) >= 3 and all(n > 50_000 for n in nums):
        is_claim_id = (
            re.match(r"^[A-Z]{2,5}[-_][A-Z]{0,3}\d{3,}", first_val, re.IGNORECASE)
            or re.match(r"^\d{4,}$", first_val.strip())
        )
        if not is_claim_id:
            return True
    return False


# ── Sheet title / metadata extractor ─────────────────────────────────────────

_LABEL_ALIASES: dict[str, str] = {
    # TPA / reinsurer
    "prepared for":     "Reinsurer",
    "reinsurer":        "Reinsurer",
    "prepared by":      "TPA Name",
    # Treaty / program
    "treaty":           "Treaty",
    "program":          "Treaty",
    "policy":           "Policy Number",
    # Cedant
    "cedant":           "Cedant",
    "ceding company":   "Cedant",
    "insurer":          "Cedant",
    # Dates
    "valuation date":   "Valuation Date",
    "valuation":        "Valuation Date",
    "as of":            "Valuation Date",
    "report date":      "Report Date",
    "report generated": "Report Date",
    "effective date":   "Effective Date",
    # Identifiers
    "policy number":    "Policy Number",
    "policy no":        "Policy Number",
    "policy #":         "Policy Number",
    "insured":          "Insured Name",
    "named insured":    "Insured Name",
    # Coverage / LOB
    "line of business": "Line of Business",
    "lob":              "Line of Business",
    "coverage":         "Coverage Type",
}


def _canonical_label(raw: str) -> str | None:
    """Map a raw label string to a canonical field name, or None if unrecognised."""
    key = raw.strip().rstrip(":").lower()
    return _LABEL_ALIASES.get(key)


def _try_inline_kv(cell_text: str) -> list[tuple[str, str]]:
    """
    Parse a single cell that contains one or more 'Key: Value' fragments.
    E.g. "Treaty: Casualty Surplus Lines 2025" or "Cedant: Hartford Financial".
    Returns a list of (raw_label, value) pairs.
    """
    pairs = []
    segments = re.split(r'\s{3,}|\|', str(cell_text))
    for seg in segments:
        m = re.match(r'^([A-Za-z][^:]{0,40}):\s*(.+)$', seg.strip())
        if m:
            pairs.append((m.group(1).strip(), m.group(2).strip()))
    return pairs


def extract_sheet_title_kvs(
    raw_rows: list,
    cell_rows: list,
    header_row_idx: int | None,
    sheet_name: str,
) -> dict:
    """
    Extract key-value metadata from the pre-header title area of a sheet.

    Scans every row above ``header_row_idx`` (or the first 15 rows when the
    sheet has no recognisable column header) and returns a dict of
    canonical-key → info-dict pairs, e.g.::

        {
            "TPA Name":       {"value": "Heritage Risk Consultants", "excel_row": 1, ...},
            "Sheet Title":    {"value": "Program Year 2025",         "excel_row": 2, ...},
            "Reinsurer":      {"value": "Munich Reinsurance …",      "excel_row": 3, ...},
            "Valuation Date": {"value": "12/31/2025",                "excel_row": 3, ...},
            "Treaty":         {"value": "Property Cat XL 2020-2025", "excel_row": 4, ...},
            "Cedant":         {"value": "Chubb Limited",             "excel_row": 4, ...},
            "Sheet Name":     {"value": "Loss Run 2025",             "excel_row": 0, ...},
        }

    Three cell patterns are handled (all without any hardcoding):

    * **Pattern A** – lone title rows (single non-empty cell, no colon):
      row 0 → ``TPA Name``; subsequent rows → ``Sheet Title``.
    * **Pattern B** – inline ``Key: Value`` in a single cell.
    * **Pattern C** – multi-cell label/value pairs on the same row.

    The sheet tab name is always stored as ``"Sheet Name"`` (source = "sheet_tab").
    """
    scan_limit = header_row_idx if header_row_idx is not None else min(15, len(raw_rows))
    found: dict = {}

    def _store(canonical: str, value: str, excel_row: int, excel_col: int):
        if canonical not in found and str(value).strip():
            found[canonical] = {
                "value":     str(value).strip(),
                "original":  str(value).strip(),
                "modified":  str(value).strip(),
                "source":    "title_kv",
                "excel_row": excel_row,
                "excel_col": excel_col,
            }

    for r_idx, row in enumerate(raw_rows[:scan_limit]):
        excel_row = r_idx + 1

        non_empty = [
            (c_idx, v) for c_idx, v in enumerate(row)
            if v is not None and str(v).strip()
        ]
        if not non_empty:
            continue

        # ── Pattern A: lone title row ─────────────────────────────────────────
        if len(non_empty) == 1:
            c_idx, val = non_empty[0]
            val_s = str(val).strip()
            if re.match(r'^[\d$,()\-\.]+$', val_s):
                continue
            if ":" not in val_s:
                if r_idx == 0:
                    tpa_name = re.split(r'\s*[\u2014\u2013]\s*', val_s)[0].strip()
                    if ' - ' in tpa_name:
                        parts = tpa_name.split(' - ', 1)
                        if re.search(r'\b(report|run|detail|summary|schedule|listing)\b',
                                     parts[1], re.IGNORECASE):
                            tpa_name = parts[0].strip()
                    _store("TPA Name", tpa_name, excel_row, c_idx + 1)
                else:
                    lob_match = re.search(
                        r'(?:loss\s+run\s+report\s*[—\-–]+\s*'
                        r'|annual\s+loss\s+run\s*[—\-–]+\s*'
                        r'|program\s+year\s+\d{4}\s*[—\-–]?\s*)(.+)',
                        val_s, re.IGNORECASE,
                    )
                    if lob_match:
                        _store("Sheet Title", lob_match.group(1).strip(), excel_row, c_idx + 1)
                    else:
                        _store("Sheet Title", val_s, excel_row, c_idx + 1)
                continue

        # ── Pattern B: inline "Key: Value" ────────────────────────────────────
        for c_idx, val in non_empty:
            val_s = str(val).strip()
            if ":" in val_s and not re.match(r'^\d', val_s):
                for raw_label, raw_value in _try_inline_kv(val_s):
                    canonical = _canonical_label(raw_label)
                    if canonical:
                        _store(canonical, raw_value, excel_row, c_idx + 1)

        # ── Pattern C: adjacent label/value cell pairs ─────────────────────────
        i = 0
        cells = non_empty
        while i < len(cells) - 1:
            c_label_idx, label_val = cells[i]
            c_value_idx, value_val = cells[i + 1]
            label_s = str(label_val).strip()
            value_s = str(value_val).strip()

            is_label = (
                label_s.endswith(":")
                or _canonical_label(label_s) is not None
            )
            if is_label and ":" in label_s and not label_s.endswith(":"):
                i += 1
                continue

            if is_label:
                canonical = (
                    _canonical_label(label_s.rstrip(":").strip())
                    or _canonical_label(label_s)
                )
                if canonical:
                    _store(canonical, value_s, excel_row, c_value_idx + 1)
                i += 2
            else:
                i += 1

    # ── Always record the sheet tab name ──────────────────────────────────────
    if "Sheet Name" not in found:
        found["Sheet Name"] = {
            "value":     sheet_name,
            "original":  sheet_name,
            "modified":  sheet_name,
            "source":    "sheet_tab",
            "excel_row": 0,
            "excel_col": 0,
        }

    return found


# ── Main entry point ──────────────────────────────────────────────────────────

def extract_from_excel(
    file_path: str,
    sheet_name: str,
) -> tuple[list, str, dict]:
    """
    Parse a single sheet from an Excel or CSV file.

    Returns
    -------
    (claims, sheet_type, title_kvs)

    ``claims``     – list of row-dicts in the standard ``{field: {value, modified, …}}``
                     format, one entry per claim row.
    ``sheet_type`` – classifier label e.g. "LOSS_RUN", "SUMMARY", "UNKNOWN".
    ``title_kvs``  – dict of canonical metadata extracted from the pre-header
                     title area (TPA Name, Treaty, Cedant, Valuation Date …).
                     Empty dict for CSV files.
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], "UNKNOWN", {}
        claims, sheet_type = parse_rows(classify_sheet(rows), rows)
        return claims, sheet_type, {}
    else:
        wb        = openpyxl.load_workbook(file_path, data_only=True)
        ws        = wb[sheet_name]
        raw_rows  = [[cell.value for cell in row] for row in ws.iter_rows()]
        cell_rows = [list(row) for row in ws.iter_rows()]
        wb.close()
        if not raw_rows:
            return [], "UNKNOWN", {}

        sheet_type = classify_sheet(raw_rows)
        hri        = _find_header_row(raw_rows)
        title_kvs  = extract_sheet_title_kvs(raw_rows, cell_rows, hri, sheet_name)
        claims, sheet_type = parse_rows_with_cells(sheet_type, raw_rows, cell_rows)
        return claims, sheet_type, title_kvs


# ── Row parsers ───────────────────────────────────────────────────────────────

def _find_header_row(rows: list) -> int | None:
    for i, row in enumerate(rows[:20]):
        rt = " ".join([str(c).lower() for c in row if c])
        if (
            "claim" in rt or "employee name" in rt or "driver name" in rt
            or "claimant" in rt or "file" in rt
        ) and (
            "date" in rt or "incurred" in rt or "paid" in rt
            or "injury" in rt or "incident" in rt or "amount" in rt or "reserve" in rt
        ):
            return i
    for i, row in enumerate(rows[:5]):
        if sum(1 for c in row if c) >= 3:
            return i
    return None


def parse_rows_with_cells(sheet_type: str, rows: list, cell_rows: list) -> tuple[list, str]:
    # ── SUMMARY sheet ─────────────────────────────────────────────────────────
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
            r_idx = hri + 2 + r_idx_rel
            if not any(raw_row):
                continue
            row_data: dict = {}
            for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
                if c_idx_0 >= len(headers):
                    continue
                clean_val = format_cell_value_with_fmt(cell)
                real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
                row_data[headers[c_idx_0]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": real_col,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    # ── Legacy print layout ───────────────────────────────────────────────────
    if _is_legacy_print_layout(rows):
        return _parse_legacy_layout_with_cells(sheet_type, rows, cell_rows)

    # ── Standard LOSS_RUN / COMMERCIAL_LOSS_RUN ───────────────────────────────
    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx_rel, (raw_row, cell_row) in enumerate(zip(rows[hri + 1:], cell_rows[hri + 1:])):
        r_idx = hri + 2 + r_idx_rel
        if not any(raw_row):
            continue
        if any(str(c).lower().strip() in ["totals", "total", "grand total", "subtotal"] for c in raw_row if c):
            break
        if _is_aggregate_row(raw_row):
            continue
        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[headers[c_idx_0]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_with_cells(
    sheet_type: str, rows: list, cell_rows: list
) -> tuple[list, str]:
    """
    Parse a legacy print-style TPA loss run sheet.

    Layout quirks handled:
    - Two-row column headers (group label row + sub-label row)
    - Address / cause-of-loss sub-rows interleaved with claim rows
    - '----------' separator rows between sections
    - 'Total …' section subtotal rows (skipped as data)

    Sub-row enrichment now uses smart pattern-based field inference via
    _enrich_from_subrow() instead of hardcoded col-B / col-D positions.
    """
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers    = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            headers = [
                str(h).strip() if h is not None else f"Column_{i}"
                for i, h in enumerate(rows[top_hri])
            ]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    num_cols = max(len(rows[i]) for i in range(len(rows))) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None

    for r_idx_rel, (raw_row, cell_row) in enumerate(
        zip(rows[data_start:], cell_rows[data_start:])
    ):
        r_idx = data_start + 1 + r_idx_rel

        # --- Completely empty row: flush pending claim -----------------------
        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Separator row (----------): flush pending, skip -----------------
        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Section subtotal row: flush pending, skip -----------------------
        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Legacy sub-row: smart field inference ---------------------------
        # Uses pattern matching to classify each cell value, with graceful
        # positional fallback for unclassifiable values (col B → Address,
        # col D → Cause of Loss).  Any extra cells stored as SubRow_<col>.
        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                _enrich_from_subrow(pending_claim, raw_row, r_idx)
            # Sub-row does NOT flush pending — the next claim row will flush it
            continue

        # --- Aggregate/totals heuristic (non-subtotal) -----------------------
        if _is_aggregate_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Normal claim row ------------------------------------------------
        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx_0, (raw_val, cell) in enumerate(zip(raw_row, cell_row)):
            if c_idx_0 >= len(headers):
                continue
            header = headers[c_idx_0]
            if not header:
                continue
            clean_val = format_cell_value_with_fmt(cell)
            real_col  = cell.column if hasattr(cell, "column") and cell.column else c_idx_0 + 1
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": real_col,
            }

        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    # Flush last pending claim
    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type


def _enrich_field(
    claim: dict, field_name: str, value: str, excel_row: int, excel_col: int
) -> None:
    """Add or update a field in a claim dict if not already set."""
    if field_name not in claim or not claim[field_name].get("value"):
        claim[field_name] = {
            "value": value, "modified": value,
            "excel_row": excel_row, "excel_col": excel_col,
        }


# ── CSV / plain parse_rows (no cell objects) ──────────────────────────────────

def parse_rows(sheet_type: str, rows: list) -> tuple[list, str]:
    if sheet_type == "SUMMARY":
        hri = None
        for i, row in enumerate(rows[:20]):
            rt = " ".join([str(c).lower() for c in row if c])
            if "sheet" in rt and "line of business" in rt:
                hri = i
                break
        if hri is None:
            return [], sheet_type
        headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        extracted = []
        for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
            if not any(row):
                continue
            if _is_aggregate_row(list(row)):
                continue
            row_data: dict = {}
            for c_idx, value in enumerate(row, start=1):
                if c_idx - 1 >= len(headers):
                    continue
                clean_val = str(value).strip() if value is not None else ""
                row_data[headers[c_idx - 1]] = {
                    "value": clean_val, "modified": clean_val,
                    "excel_row": r_idx, "excel_col": c_idx,
                }
            if any(v["value"] for v in row_data.values()):
                extracted.append(row_data)
        return extracted, sheet_type

    if _is_legacy_print_layout(rows):
        return _parse_legacy_layout_plain(sheet_type, rows)

    hri = _find_header_row(rows)
    if hri is None:
        return [], sheet_type
    headers   = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
    extracted = []
    for r_idx, row in enumerate(rows[hri + 1:], start=hri + 2):
        if not any(row):
            continue
        if any(str(cell).lower().strip() in ["totals", "total", "grand total"] for cell in row if cell):
            break
        if _is_aggregate_row(list(row)):
            continue
        row_data: dict = {}
        for c_idx, value in enumerate(row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[headers[c_idx - 1]] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            extracted.append(row_data)
    return extracted, sheet_type


def _parse_legacy_layout_plain(sheet_type: str, rows: list) -> tuple[list, str]:
    """
    parse_rows equivalent for legacy layout when no cell objects are available.

    Sub-row enrichment uses the same smart _enrich_from_subrow() as the
    cell-aware version.
    """
    header_pair = _find_legacy_header_rows(rows)
    if header_pair is None:
        hri = _find_header_row(rows)
        if hri is None:
            return [], sheet_type
        headers    = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[hri])]
        data_start = hri + 1
    else:
        top_hri, bot_hri = header_pair
        if top_hri == bot_hri:
            headers = [str(h).strip() if h is not None else f"Column_{i}" for i, h in enumerate(rows[top_hri])]
        else:
            headers = _merge_two_header_rows(rows[top_hri], rows[bot_hri])
        data_start = bot_hri + 1

    num_cols = max(len(r) for r in rows) if rows else len(headers)
    while len(headers) < num_cols:
        headers.append(f"Column_{len(headers) + 1}")

    extracted: list[dict] = []
    pending_claim: dict | None = None

    for r_idx, raw_row in enumerate(rows[data_start:], start=data_start + 1):
        if not any(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_separator_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue
        if _is_subtotal_row(raw_row):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        # --- Legacy sub-row: smart field inference ---------------------------
        if _is_legacy_sub_row(raw_row, num_cols):
            if pending_claim is not None:
                _enrich_from_subrow(pending_claim, raw_row, r_idx)
            continue

        if _is_aggregate_row(list(raw_row)):
            if pending_claim is not None:
                extracted.append(pending_claim)
                pending_claim = None
            continue

        if pending_claim is not None:
            extracted.append(pending_claim)
            pending_claim = None

        row_data: dict = {}
        for c_idx, value in enumerate(raw_row, start=1):
            if c_idx - 1 >= len(headers):
                continue
            header    = headers[c_idx - 1]
            if not header:
                continue
            clean_val = str(value).strip() if value is not None else ""
            row_data[header] = {
                "value": clean_val, "modified": clean_val,
                "excel_row": r_idx, "excel_col": c_idx,
            }
        if any(v["value"] for v in row_data.values()):
            pending_claim = row_data

    if pending_claim is not None:
        extracted.append(pending_claim)

    return extracted, sheet_type
