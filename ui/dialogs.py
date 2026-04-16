"""
ui/dialogs.py  — light-theme version
All @st.dialog popups with white backgrounds and dark text.
"""

import csv
import datetime

import streamlit as st
from openpyxl.utils import get_column_letter
from PIL import ImageDraw

from modules.field_history import _get_field_history
from modules.excel_renderer import (
    render_excel_sheet, get_cell_pixel_bbox, crop_context,
)

# ── Light theme color constants ───────────────────────────────────────────────
_D_BG   = "#f8f9fc"   # card background
_D_BG2  = "#f1f3f8"   # deeper background
_D_BDR  = "#d0d6e8"   # border
_D_BDR2 = "#e8ecf4"   # lighter border
_D_TXT  = "#0f1117"   # main text
_D_LBL  = "#4a5578"   # subdued label
_D_GRN  = "#0a9e6a"   # green
_D_GRN_BG = "#e6f9f2"
_D_RED  = "#d64040"
_D_RED_BG = "#fff0f0"
_D_YEL  = "#c99a00"
_D_YEL_BG = "#fffbeb"
_D_BLU  = "#1a6fd8"
_D_BLU_BG = "#e8f0fe"
_D_PUR  = "#6b3fd4"


# ── Eye popup ─────────────────────────────────────────────────────────────────

@st.dialog("Cell View", width="large")
def show_eye_popup(field: str, info: dict, excel_path: str, sheet_name: str) -> None:
    import os
    raw_value  = info.get("value", "") or ""
    mod_value  = info.get("modified", raw_value) or raw_value
    target_row = info.get("excel_row")
    target_col = info.get("excel_col")

    st.markdown(f"### 📍 {field}")

    def _val_box(label: str, val: str, color: str = _D_BLU):
        _content = val if val else f"<span style='color:{_D_LBL};'>( empty )</span>"
        st.markdown(
            f"<div style='margin-bottom:12px;'>"
            f"<div style='font-size:10px;font-weight:700;color:{color};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1.2px;margin-bottom:5px;'>{label}</div>"
            f"<div style='background:{_D_BG2};border:1px solid {_D_BDR};border-radius:6px;"
            f"padding:12px 14px;font-family:Consolas,monospace;font-size:13px;"
            f"color:{_D_TXT};word-break:break-all;white-space:pre-wrap;"
            f"max-height:200px;overflow-y:auto;line-height:1.6;'>"
            f"{_content}"
            f"</div></div>",
            unsafe_allow_html=True,
        )

    _val_box("Extracted Value (raw from file)", raw_value, _D_GRN)
    if mod_value and mod_value != raw_value:
        _val_box("Modified Value (user edited)", mod_value, _D_YEL)

    ext = os.path.splitext(excel_path)[1].lower()

    if target_row and target_col:
        col_letter = get_column_letter(target_col)
        st.markdown(
            f"<div style='font-size:12px;color:{_D_BLU};font-family:monospace;"
            f"font-weight:600;margin-bottom:12px;'>"
            f"📌 Cell <b>{col_letter}{target_row}</b>"
            f" &nbsp;·&nbsp; Row {target_row} &nbsp;·&nbsp; Col {target_col}"
            f"</div>",
            unsafe_allow_html=True,
        )
    elif target_row and ext == ".pdf":
        st.markdown(
            f"<div style='font-size:12px;color:{_D_BLU};font-family:monospace;"
            f"font-weight:600;margin-bottom:12px;'>"
            f"📄 PDF Page <b>{target_row}</b>"
            f"</div>",
            unsafe_allow_html=True,
        )
    else:
        st.warning("No cell location recorded for this field.")
        return

    st.markdown("---")

    # ── PDF branch ────────────────────────────────────────────────────────────
    if ext == ".pdf":
        source_text      = info.get("source_text", "")
        bounding_polygon = info.get("bounding_polygon")
        page_width       = info.get("page_width")  or 8.5
        page_height      = info.get("page_height") or 11.0

        st.markdown(
            f"<div style='font-size:10px;color:{_D_BLU};font-weight:700;font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;'>"
            f"📄 PDF Source — Page {target_row}</div>",
            unsafe_allow_html=True,
        )

        _src_box_style = (
            f"background:{_D_BG2};border:1px solid {_D_BDR};border-radius:6px;"
            f"padding:8px 12px;font-family:monospace;font-size:11px;color:{_D_TXT};margin-top:8px;"
        )

        if bounding_polygon:
            _pdf_cache_key = f"_pdf_render_{excel_path}_{target_row}_{field}"
            with st.spinner("Rendering PDF page…"):
                if _pdf_cache_key not in st.session_state:
                    try:
                        from modules.excel_renderer import render_pdf_page_with_highlight
                        full_img, cropped_img = render_pdf_page_with_highlight(
                            pdf_path=excel_path, page_number=int(target_row),
                            bounding_polygon=bounding_polygon,
                            page_width_inches=float(page_width),
                            page_height_inches=float(page_height), dpi=150,
                        )
                        st.session_state[_pdf_cache_key] = (full_img, cropped_img)
                    except Exception as e:
                        st.session_state[_pdf_cache_key] = (None, None)
                        st.error(f"PDF render error: {e}")
                full_img, cropped_img = st.session_state.get(_pdf_cache_key, (None, None))
            if cropped_img is not None:
                st.image(cropped_img, use_container_width=True,
                         caption=f"Field '{field}' highlighted on PDF Page {target_row}")
                if source_text:
                    st.markdown(f"<div style='{_src_box_style}'>{source_text}</div>",
                                unsafe_allow_html=True)
            else:
                st.warning("Could not render PDF page image. Install pymupdf: `pip install pymupdf`")
                st.markdown(
                    f"<div style='{_src_box_style}'>{source_text or '(no source text recorded)'}</div>",
                    unsafe_allow_html=True,
                )
        else:
            _pdf_cache_key = f"_pdf_render_{excel_path}_{target_row}_{field}_textfallback"
            with st.spinner("Rendering PDF page…"):
                if _pdf_cache_key not in st.session_state:
                    try:
                        from modules.excel_renderer import render_pdf_page_text_highlight
                        full_img, cropped_img = render_pdf_page_text_highlight(
                            pdf_path=excel_path, page_number=int(target_row),
                            search_text=raw_value or source_text or field, dpi=150,
                        )
                        st.session_state[_pdf_cache_key] = (full_img, cropped_img)
                    except Exception as e:
                        st.session_state[_pdf_cache_key] = (None, None)
                full_img, cropped_img = st.session_state.get(_pdf_cache_key, (None, None))
            if cropped_img is not None:
                st.image(cropped_img, use_container_width=True,
                         caption=f"Field '{field}' highlighted on PDF Page {target_row}")
            else:
                st.markdown(
                    f"<div style='{_src_box_style}line-height:1.6;margin-bottom:8px;'>"
                    f"{source_text or '(no source text recorded)'}</div>",
                    unsafe_allow_html=True,
                )
            if source_text:
                st.markdown(f"<div style='{_src_box_style}'>{source_text}</div>",
                            unsafe_allow_html=True)
            st.info("📝 This field was extracted from page text — highlighted by text search.")
        return

    # ── CSV branch ────────────────────────────────────────────────────────────
    if ext == ".csv":
        try:
            with open(excel_path, "r", encoding="utf-8-sig") as f:
                all_rows = list(csv.reader(f))
            if not all_rows:
                return
            n_rows = len(all_rows)
            n_cols = max(len(r) for r in all_rows)
            r0, r1 = max(0, target_row - 4), min(n_rows, target_row + 4)

            col_headers = "".join(
                f"<th style='background:{_D_BG2};color:{_D_TXT};font-size:11px;"
                f"padding:5px 10px;border:1px solid {_D_BDR};font-family:monospace;"
                f"font-weight:700;text-align:center;'>{get_column_letter(c+1)}</th>"
                for c in range(n_cols)
            )
            thead = (
                f"<thead><tr>"
                f"<th style='background:{_D_BG2};color:{_D_TXT};font-size:11px;"
                f"padding:5px 8px;border:1px solid {_D_BDR};font-family:monospace;font-weight:700;'>#</th>"
                f"{col_headers}</tr></thead>"
            )
            tbody = ""
            for r_idx in range(r0, r1):
                row_data = all_rows[r_idx] if r_idx < len(all_rows) else []
                is_tr    = (r_idx + 1 == target_row)
                rn_bg    = _D_BLU_BG if is_tr else "#ffffff"
                rn_color = _D_BLU if is_tr else _D_LBL
                cells = (
                    f"<td style='background:{rn_bg};color:{rn_color};font-size:11px;"
                    f"padding:5px 8px;border:1px solid {_D_BDR};font-family:monospace;"
                    f"font-weight:bold;text-align:center;'>{r_idx+1}</td>"
                )
                for c_idx in range(n_cols):
                    cell_val = row_data[c_idx] if c_idx < len(row_data) else ""
                    is_tc    = is_tr and (c_idx + 1 == target_col)
                    if is_tc:
                        style = f"background:{_D_YEL_BG};border:2px solid {_D_YEL};color:{_D_TXT};font-weight:bold;"
                    elif is_tr:
                        style = f"background:{_D_BLU_BG};border:1px solid #b3c8f5;color:{_D_TXT};"
                    else:
                        style = f"background:#ffffff;border:1px solid {_D_BDR};color:{_D_TXT};"
                    cells += (
                        f"<td style='{style}font-size:11px;padding:5px 10px;"
                        f"white-space:normal;word-break:break-word;font-family:monospace;'>{cell_val}</td>"
                    )
                tbody += f"<tr>{cells}</tr>"

            st.markdown(
                f"<div style='overflow-x:auto;border-radius:6px;border:1px solid {_D_BDR};'>"
                f"<table style='border-collapse:collapse;width:100%;'>"
                f"{thead}<tbody>{tbody}</tbody></table></div>",
                unsafe_allow_html=True,
            )
        except Exception as e:
            st.error(f"CSV preview error: {e}")
        return

    # ── Excel branch ──────────────────────────────────────────────────────────
    cache_key = f"_rendered_{excel_path}_{sheet_name}"
    with st.spinner("Rendering sheet…"):
        if cache_key not in st.session_state:
            rendered_img, col_starts, row_starts, merged_master = render_excel_sheet(
                excel_path, sheet_name, scale=1.0
            )
            st.session_state[cache_key] = (rendered_img, col_starts, row_starts, merged_master)
        else:
            rendered_img, col_starts, row_starts, merged_master = st.session_state[cache_key]

    try:
        img  = rendered_img.copy()
        draw = ImageDraw.Draw(img, "RGBA")
        x1, y1, x2, y2 = get_cell_pixel_bbox(col_starts, row_starts, target_row, target_col, merged_master)
        draw.rectangle([x1+1, y1+1, x2-1, y2-1], fill=(255, 230, 0, 80))
        draw.rectangle([x1, y1, x2, y2], outline=(245, 158, 11, 255), width=3)
        draw.rectangle([x1+3, y1+3, x2-3, y2-3], outline=(255, 255, 255, 160), width=1)
        cropped, _, _, _, _ = crop_context(img, x1, y1, x2, y2, pad_x=300, pad_y=200)
        col_letter = get_column_letter(target_col)
        st.image(cropped, use_container_width=True,
                 caption=f"Cell {col_letter}{target_row} highlighted in yellow")
    except Exception as e:
        st.error(f"Rendering error: {e}")


# ── Field history dialog ──────────────────────────────────────────────────────

@st.dialog("Field History", width="large")
def show_field_history_dialog(
    field_name: str, sheet: str, claim_id: str,
    current_val: str, original_val: str,
) -> None:
    st.markdown(f"### 📋 History — {field_name}")
    history = _get_field_history(sheet, claim_id, field_name)

    orig_style = f"background:{_D_BG2};border:1px solid {_D_BDR};border-radius:5px;padding:8px 12px;font-family:monospace;font-size:13px;color:{_D_TXT};"
    curr_style = f"background:{_D_GRN_BG};border:1px solid {_D_GRN}60;border-radius:5px;padding:8px 12px;font-family:monospace;font-size:13px;color:{_D_GRN};"
    mod_note = (
        f"<div style='margin-top:8px;font-size:11px;color:{_D_YEL};font-family:monospace;'>⚡ Modified from original</div>"
        if current_val != original_val else
        f"<div style='margin-top:8px;font-size:11px;color:{_D_GRN};font-family:monospace;'>✓ Unchanged from original</div>"
    )

    st.markdown(
        f"<div style='background:{_D_BG};border:1px solid {_D_BDR};border-radius:8px;"
        f"padding:12px 16px;margin-bottom:12px;'>"
        f"<div style='display:grid;grid-template-columns:1fr 1fr;gap:16px;'>"
        f"<div>"
        f"<div style='font-size:10px;color:{_D_LBL};font-family:monospace;text-transform:uppercase;"
        f"letter-spacing:1px;margin-bottom:6px;'>Original (from file)</div>"
        f"<div style='{orig_style}'>{original_val or '(empty)'}</div>"
        f"</div>"
        f"<div>"
        f"<div style='font-size:10px;color:{_D_LBL};font-family:monospace;text-transform:uppercase;"
        f"letter-spacing:1px;margin-bottom:6px;'>Current Value</div>"
        f"<div style='{curr_style}'>{current_val or '(empty)'}</div>"
        f"</div></div>"
        f"{mod_note}"
        f"</div>",
        unsafe_allow_html=True,
    )

    if history:
        st.markdown("**Edit Timeline**")
        for h in history:
            arrow_col = _D_YEL if h["source"] == "user" else _D_BLU
            src_icon  = "✏" if h["source"] == "user" else "⚡"
            src_lbl   = "Manual edit" if h["source"] == "user" else "Auto (LLM/normalize)"
            from_style = f"background:{_D_BG2};padding:3px 8px;border-radius:4px;font-size:12px;color:{_D_TXT};border:1px solid {_D_BDR};"
            to_style   = f"background:{_D_GRN_BG};padding:3px 8px;border-radius:4px;font-size:12px;color:{_D_GRN};"
            st.markdown(
                f"<div style='display:flex;align-items:flex-start;gap:12px;padding:10px 0;"
                f"border-bottom:1px solid {_D_BDR2};'>"
                f"<div style='font-size:10px;color:{_D_LBL};font-family:monospace;"
                f"white-space:nowrap;margin-top:2px;'>{h['ts']}</div>"
                f"<div style='flex:1;'>"
                f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:6px;'>"
                f"<span style='color:{arrow_col};font-size:12px;'>{src_icon}</span>"
                f"<span style='font-size:11px;color:{_D_LBL};font-family:monospace;'>{src_lbl}</span></div>"
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<code style='{from_style}'>{h['from'] or '(empty)'}</code>"
                f"<span style='color:{arrow_col};font-size:14px;'>→</span>"
                f"<code style='{to_style}'>{h['to'] or '(empty)'}</code>"
                f"</div></div></div>",
                unsafe_allow_html=True,
            )
    else:
        st.markdown(
            f"<div style='color:{_D_LBL};font-size:13px;padding:12px 0;'>"
            "No edits recorded yet for this field.</div>",
            unsafe_allow_html=True,
        )

    if st.button("Close", type="primary", use_container_width=True):
        st.rerun()


# ── Settings dialog ───────────────────────────────────────────────────────────

@st.dialog("Settings", width="large")
def show_settings_dialog(schemas: dict, config_load_status: dict) -> None:
    import os
    from config.settings import CONFIG_DIR

    st.markdown("### Configuration")
    st.markdown("---")
    st.markdown("#### Confidence Settings")

    use_conf = st.checkbox(
        "Enable confidence scoring display",
        value=st.session_state.get("use_conf_threshold", False),
        key="use_conf_toggle",
        help="When enabled, shows confidence scores for each mapped field",
    )
    st.session_state["use_conf_threshold"] = use_conf

    if use_conf:
        conf = st.slider(
            "Confidence threshold", 0, 100,
            value=st.session_state.get("conf_threshold", 80),
            step=5, format="%d%%",
        )
        st.session_state["conf_threshold"] = conf
        bar_color = "#22c55e" if conf >= 70 else "#f59e0b" if conf >= 40 else "#ef4444"
        level_txt = (
            "High confidence — minimal manual review needed" if conf >= 70 else
            "Medium — review flagged fields carefully" if conf >= 40 else
            "Low — most fields will require manual review"
        )
        st.markdown(
            f"<div style='background:{_D_BG2};border-radius:4px;height:5px;width:100%;margin-top:4px;overflow:hidden;'>"
            f"<div style='width:{conf}%;background:{bar_color};height:100%;border-radius:4px;'></div></div>"
            f"<div style='color:{bar_color};font-size:12px;margin-top:5px;'>{level_txt}</div>",
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f"<div style='color:{_D_LBL};font-size:13px;font-family:monospace;'>"
            "Confidence scoring is disabled. Enable above to show scores and set threshold.</div>",
            unsafe_allow_html=True,
        )

    st.markdown("---")
    st.markdown("#### Export Schema")
    active_schema = st.session_state.get("active_schema", None)

    for schema_name, schema_def in schemas.items():
        is_active  = active_schema == schema_name
        border_col = schema_def["color"] if is_active else _D_BDR
        bg_col     = _D_BLU_BG if is_active else _D_BG
        active_tag = (
            f"<span style='font-size:10px;color:{schema_def['color']};margin-left:8px;"
            f"font-weight:bold;'>● ACTIVE</span>"
            if is_active else ""
        )
        custom_count = len(st.session_state.get(f"custom_fields_{schema_name}", []))
        st.markdown(
            f"<div style='background:{bg_col};border:1px solid {border_col};border-radius:8px;"
            f"padding:12px 14px;margin-bottom:4px;'>"
            f"<div style='display:flex;align-items:center;'>"
            f"<span style='font-size:14px;font-weight:700;color:{_D_TXT};font-family:var(--font);'>"
            f"{schema_def['icon']} {schema_name}</span>"
            f"<span style='font-size:13px;color:{_D_LBL};margin-left:8px;font-family:var(--font);'>"
            f"{schema_def['version']}</span>{active_tag}</div>"
            f"<div style='font-size:13px;color:{_D_LBL};margin-top:4px;font-family:var(--font);'>"
            f"{schema_def['description']}</div></div>",
            unsafe_allow_html=True,
        )
        bc1, bc2, bc3 = st.columns([1, 1, 1])
        with bc1:
            if st.button(
                "✓ Deactivate" if is_active else "Activate",
                key=f"activate_{schema_name}", use_container_width=True,
            ):
                st.session_state["active_schema"] = None if is_active else schema_name
                st.rerun()
        with bc2:
            if st.button("View Fields", key=f"view_{schema_name}", use_container_width=True):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "required"
                st.rerun()
        with bc3:
            if st.button(
                f"Custom Fields ({custom_count})",
                key=f"custom_{schema_name}", use_container_width=True,
            ):
                st.session_state["schema_popup_target"] = schema_name
                st.session_state["schema_popup_tab"]    = "custom"
                st.rerun()
        st.markdown("<div style='height:6px;'></div>", unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("#### 📁 YAML Config Files")
    st.markdown(
        f"<div style='color:{_D_LBL};font-size:14px;margin-bottom:10px;font-family:var(--font);'>"
        f"Config directory: <code>{CONFIG_DIR}</code></div>",
        unsafe_allow_html=True,
    )
    for schema_name, status in config_load_status.items():
        sc     = schemas.get(schema_name, {})
        col_st = sc.get("color", "#64748b")
        badge  = (
            f"<span style='background:{_D_GRN_BG};border:1px solid {_D_GRN};border-radius:4px;"
            f"padding:1px 7px;font-size:10px;color:{_D_GRN};font-weight:600;'>✓ Loaded</span>"
            if status["loaded"]
            else
            f"<span style='background:{_D_RED_BG};border:1px solid {_D_RED};border-radius:4px;"
            f"padding:1px 7px;font-size:10px;color:{_D_RED};font-weight:600;'>✗ Not found — using defaults</span>"
        )
        st.markdown(
            f"<div style='background:{_D_BG};border:1px solid {_D_BDR};border-radius:6px;"
            f"padding:10px 14px;margin-bottom:6px;'>"
            f"<div style='display:flex;align-items:center;gap:10px;'>"
            f"<span style='color:{col_st};font-weight:700;font-size:14px;font-family:var(--font);'>"
            f"{sc.get('icon','')} {schema_name}</span>{badge}</div>"
            f"<div style='font-size:12px;color:{_D_LBL};margin-top:4px;font-family:var(--font);'>"
            f"📄 {status['file']}</div></div>",
            unsafe_allow_html=True,
        )

    if st.button("🔄 Reload YAML Configs", use_container_width=True, key="reload_yaml_cfg"):
        from config.schemas import _load_all_configs, _HARDCODED_SCHEMAS
        import config.schemas as _cs
        _cs.SCHEMAS = _load_all_configs(_HARDCODED_SCHEMAS)
        st.session_state["sheet_cache"] = {}
        st.success("✅ Configs reloaded")
        st.rerun()

    st.markdown("---")
    r1, r2 = st.columns(2)
    with r1:
        if st.button("Reset Defaults", use_container_width=True, key="reset_defaults_btn"):
            st.session_state["conf_threshold"]     = 80
            st.session_state["use_conf_threshold"] = False
            st.session_state["active_schema"]      = None
            for s in schemas:
                st.session_state[f"custom_fields_{s}"] = []
            st.rerun()
    with r2:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


# ── Schema field manager dialog ───────────────────────────────────────────────

@st.dialog("Schema Field Manager", width="large")
def show_schema_fields_dialog(schema_name: str, schemas: dict) -> None:
    PILL_STYLE_REQ    = (
        "display:inline-block;background:#e8f0fe;border:1px solid #1a6fd860;border-radius:6px;padding:5px 12px;font-size:12px;color:#1a6fd8;font-weight:600;margin:4px 5px 4px 0;font-family:monospace;white-space:nowrap;"
    )
    PILL_STYLE_OPT    = (
        "display:inline-block;background:#f1f3f8;border:1px solid #d0d6e8;border-radius:6px;padding:5px 12px;font-size:12px;color:#0f1117;font-weight:500;margin:4px 5px 4px 0;font-family:monospace;white-space:nowrap;"
    )
    PILL_STYLE_CUSTOM = (
        "display:inline-block;background:#e6f9f2;border:1px solid #0a9e6a60;border-radius:6px;padding:5px 12px;font-size:12px;color:#0a9e6a;font-weight:600;margin:4px 5px 4px 0;font-family:monospace;white-space:nowrap;"
    )
    schema     = schemas[schema_name]
    custom_key = f"custom_fields_{schema_name}"
    if custom_key not in st.session_state:
        st.session_state[custom_key] = []

    st.markdown(f"### {schema['icon']} {schema_name} — {schema['version']}")
    st.markdown(
        f"<div style='color:{_D_LBL};font-size:14px;margin-bottom:14px;font-family:var(--font);'>"
        f"{schema['description']}</div>",
        unsafe_allow_html=True,
    )
    tab_req, tab_accepted, tab_custom = st.tabs(["Mandatory Fields", "All Accepted Fields", "My Custom Fields"])

    with tab_req:
        pills = "".join(
            f"<span style='display:inline-block;background:#e8f0fe;border:1px solid #1a6fd860;border-radius:6px;padding:5px 12px;font-size:12px;color:#1a6fd8;font-weight:600;margin:4px 5px 4px 0;font-family:monospace;white-space:nowrap;'>✓ {f}</span>"
            for f in schema["required_fields"]
        )
        st.markdown(f"<div style='display:flex;flex-wrap:wrap;gap:0;margin:12px 0;'>{pills}</div>", unsafe_allow_html=True)

    with tab_accepted:
        optional  = [f for f in schema["accepted_fields"] if f not in schema["required_fields"]]
        req_pills = "".join(f"<span style='display:inline-block;background:#e8f0fe;border:1px solid #1a6fd860;border-radius:6px;padding:5px 12px;font-size:12px;color:#1a6fd8;font-weight:600;margin:4px 5px 4px 0;font-family:monospace;white-space:nowrap;'>✓ {f}</span>" for f in schema["required_fields"])
        opt_pills = "".join(f"<span style='display:inline-block;background:#f1f3f8;border:1px solid #d0d6e8;border-radius:6px;padding:5px 12px;font-size:12px;color:#0f1117;font-weight:500;margin:4px 5px 4px 0;font-family:monospace;white-space:nowrap;'>{f}</span>" for f in optional)
        st.markdown(
            f"<div style='margin-bottom:16px;'>"
            f"<div style='font-size:10px;font-weight:800;color:{_D_TXT};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1.5px;margin-bottom:8px;"
            f"padding-bottom:4px;border-bottom:2px solid {_D_BDR};'>✓ Mandatory Fields</div>"
            f"<div style='display:flex;flex-wrap:wrap;gap:0;'>{req_pills}</div></div>"
            f"<div style='margin-bottom:16px;'>"
            f"<div style='font-size:10px;font-weight:800;color:{_D_TXT};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1.5px;margin-bottom:8px;"
            f"padding-bottom:4px;border-bottom:2px solid {_D_BDR};'>Optional Fields</div>"
            f"<div style='display:flex;flex-wrap:wrap;gap:0;'>{opt_pills}</div></div>",
            unsafe_allow_html=True,
        )

    with tab_custom:
        custom_fields = st.session_state[custom_key]
        already_added = set(custom_fields) | set(schema["required_fields"])
        available     = [f for f in schema["accepted_fields"] if f not in already_added]

        if available:
            sel_col, add_col = st.columns([4, 1])
            with sel_col:
                chosen = st.selectbox(
                    "Pick field", ["— select a field —"] + available,
                    key=f"new_field_sel_{schema_name}", label_visibility="collapsed",
                )
            with add_col:
                if st.button("Add", key=f"add_field_btn_{schema_name}", use_container_width=True, type="primary"):
                    if chosen and chosen != "— select a field —":
                        st.session_state[custom_key].append(chosen)
                        st.rerun()

        if not custom_fields:
            st.markdown(
                f"<div style='color:{_D_LBL};font-size:14px;padding:10px 0;font-family:var(--font);'>"
                "No optional fields added yet.</div>",
                unsafe_allow_html=True,
            )
        else:
            for idx, cf in enumerate(list(custom_fields)):
                cf1, cf2 = st.columns([5, 1])
                with cf1:
                    p_style = PILL_STYLE_REQ if cf in schema["required_fields"] else PILL_STYLE_CUSTOM
                    icon    = "✓" if cf in schema["required_fields"] else "+"
                    st.markdown(f"<span style='{p_style}'>{icon} {cf}</span>", unsafe_allow_html=True)
                with cf2:
                    if st.button("Remove", key=f"del_cf_{schema_name}_{idx}", use_container_width=True):
                        st.session_state[custom_key].pop(idx)
                        st.rerun()
            st.markdown("---")
            if st.button(f"Clear All", key=f"clear_all_{schema_name}"):
                st.session_state[custom_key] = []
                st.rerun()

        total = len(schema["required_fields"]) + len(custom_fields)
        st.markdown(
            f"<div style='background:{_D_BG};border:1px solid {_D_BDR};border-radius:8px;"
            f"padding:10px 16px;'>"
            f"<span style='color:{_D_LBL};font-size:14px;font-family:var(--font);'>"
            f"Mandatory: <b style='color:{_D_BLU};'>{len(schema['required_fields'])}</b>"
            f" &nbsp;|&nbsp; Custom: <b style='color:{_D_GRN};'>{len(custom_fields)}</b>"
            f" &nbsp;|&nbsp; Total: <b style='color:{_D_TXT};'>{total}</b></span></div>",
            unsafe_allow_html=True,
        )


# ── Cache Manager dialog ──────────────────────────────────────────────────────

@st.dialog("Cache Manager", width="large")
def show_cache_manager_dialog() -> None:
    from modules.cache_manager import (
        get_cache_stats, clear_session_cache, clear_parsed_cache,
        clear_hash_store, clear_claim_dup_store,
        clear_audit_log, clear_export_table, _fmt_size,
    )

    st.markdown("### 🗄️ Cache Manager")
    st.markdown(
        f"<div style='font-size:13px;color:{_D_LBL};margin-bottom:16px;'>"
        "View and selectively clear each cache layer. This does not affect your uploaded files.</div>",
        unsafe_allow_html=True,
    )

    stats = get_cache_stats()

    def _stat_row(label, detail, color=_D_BLU):
        st.markdown(
            f"<div style='display:flex;justify-content:space-between;align-items:center;"
            f"background:{_D_BG};border:1px solid {_D_BDR};border-radius:6px;"
            f"padding:10px 14px;margin-bottom:6px;'>"
            f"<div style='font-size:13px;font-weight:600;color:{_D_TXT};'>{label}</div>"
            f"<div style='font-size:12px;font-family:monospace;color:{color};font-weight:600;'>{detail}</div>"
            f"</div>",
            unsafe_allow_html=True,
        )

    _stat_row("Parsed Sheet Cache",
              f"{stats['parsed']['files']} file(s) · {_fmt_size(stats['parsed']['size_kb'])}",
              _D_GRN if stats['parsed']['files'] > 0 else _D_LBL)
    _stat_row("File Hash Store (Duplicate Memory)",
              f"{stats['hash_store']['entries']} file(s) tracked",
              _D_YEL if stats['hash_store']['entries'] > 0 else _D_LBL)
    _stat_row("Claim Duplicate Store",
              f"{stats['claim_dups']['entries']} claim(s) tracked",
              _D_RED if stats['claim_dups']['entries'] > 0 else _D_LBL)
    _stat_row("Audit Log",
              f"{stats['audit_log']['entries']} event(s) recorded",
              _D_PUR if stats['audit_log']['entries'] > 0 else _D_LBL)
    _stat_row("Export History",
              f"{stats['export_table']['entries']} export(s) recorded",
              _D_BLU if stats['export_table']['entries'] > 0 else _D_LBL)

    st.markdown("---")
    st.markdown(
        f"<div style='font-size:12px;color:{_D_TXT};font-weight:700;margin-bottom:10px;font-family:monospace;"
        f"text-transform:uppercase;letter-spacing:1px;'>Select what to clear</div>",
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        do_session   = st.checkbox("UI Session State", value=True)
        do_parsed    = st.checkbox("Parsed Sheet Cache", value=False)
        do_hash      = st.checkbox("File Duplicate Memory", value=False)
    with c2:
        do_claim_dup = st.checkbox("Claim Duplicate Store", value=False)
        do_audit     = st.checkbox("Audit Log", value=False)
        do_exports   = st.checkbox("Export History", value=False)

    st.markdown("---")
    st.markdown(
        f"<div style='font-size:12px;color:{_D_TXT};font-weight:700;margin-bottom:8px;font-family:monospace;"
        f"text-transform:uppercase;letter-spacing:1px;'>Quick presets</div>",
        unsafe_allow_html=True,
    )
    p1, p2, p3 = st.columns(3)
    with p1:
        if st.button("🔄 Soft Reset", use_container_width=True):
            cleared = clear_session_cache(st.session_state)
            st.success(f"✅ UI state cleared ({cleared} keys removed)")
            st.rerun()
    with p2:
        if st.button("📁 Clear File History", use_container_width=True):
            n_hash = clear_hash_store()
            n_dup  = clear_claim_dup_store()
            st.session_state["sheet_cache"] = {}
            st.success(f"✅ File history cleared ({n_hash} files, {n_dup} claims reset)")
            st.rerun()
    with p3:
        if st.button("🗑️ Full Reset", use_container_width=True, type="primary"):
            st.session_state["_confirm_full_reset"] = True
            st.rerun()

    if st.session_state.get("_confirm_full_reset"):
        st.warning("⚠️ **This will clear ALL cache layers.** Are you sure?")
        yes_col, no_col = st.columns(2)
        with yes_col:
            if st.button("Yes, clear everything", type="primary", use_container_width=True):
                clear_session_cache(st.session_state)
                clear_parsed_cache()
                clear_hash_store()
                clear_claim_dup_store()
                clear_audit_log()
                clear_export_table()
                st.session_state["_confirm_full_reset"] = False
                st.success("✅ All cache layers cleared.")
                st.rerun()
        with no_col:
            if st.button("Cancel", use_container_width=True):
                st.session_state["_confirm_full_reset"] = False
                st.rerun()

    st.markdown("---")
    col_clear, col_close = st.columns(2)
    with col_clear:
        if st.button("🗑️ Clear Selected", use_container_width=True):
            msgs = []
            if do_session:
                n = clear_session_cache(st.session_state)
                msgs.append(f"UI state ({n} keys)")
            if do_parsed:
                files, kb = clear_parsed_cache()
                msgs.append(f"Parsed cache ({files} files, {_fmt_size(kb)})")
                st.session_state["sheet_cache"] = {}
            if do_hash:
                n = clear_hash_store()
                msgs.append(f"File history ({n} entries)")
            if do_claim_dup:
                n = clear_claim_dup_store()
                msgs.append(f"Claim dups ({n} entries)")
            if do_audit:
                n = clear_audit_log()
                msgs.append(f"Audit log ({n} events)")
            if do_exports:
                n = clear_export_table()
                msgs.append(f"Export history ({n} entries)")
            if msgs:
                st.success("✅ Cleared: " + ", ".join(msgs))
                st.rerun()
            else:
                st.warning("Nothing selected — tick at least one checkbox above.")
    with col_close:
        if st.button("Close", type="primary", use_container_width=True):
            st.rerun()


# ── Claim Journey / Traceability Dialog ───────────────────────────────────────

@st.dialog("Claim Transformation Journey", width="large")
def show_claim_journey_dialog(
    claim_id: str,
    curr_claim: dict,
    selected_sheet: str,
    active_schema: str | None,
    _llm_map_result: dict,
) -> None:
    import json as _json
    import datetime as _dt
    from modules.audit import _load_audit_log
    from modules.field_history import _get_field_history
    from modules.schema_mapping import map_claim_to_schema

    _audit_expand_key = f"_audit_expanded_{selected_sheet}_{claim_id}"
    _full_hist_key    = f"_audit_fullhist_{selected_sheet}_{claim_id}"
    if _audit_expand_key not in st.session_state:
        st.session_state[_audit_expand_key] = set()
    if _full_hist_key not in st.session_state:
        st.session_state[_full_hist_key] = False

    _ts_dialog_open = _dt.datetime.now()
    _ts_fmt = lambda d: d.strftime("%H:%M:%S.%f")[:-3]

    st.markdown(
        f"<div style='font-size:18px;font-weight:700;color:{_D_TXT};margin-bottom:4px;'>"
        f"🔍 Transformation Journey</div>"
        f"<div style='font-size:12px;color:{_D_LBL};font-family:monospace;margin-bottom:4px;font-weight:600;'>"
        f"Claim {claim_id} · Sheet: {selected_sheet}"
        + (f" · Schema: {active_schema}" if active_schema else "")
        + f"</div>"
        f"<div style='font-size:10px;color:{_D_LBL};font-family:monospace;margin-bottom:16px;'>"
        f"⏱ Dialog opened at {_ts_fmt(_ts_dialog_open)}</div>",
        unsafe_allow_html=True,
    )

    _all_audit   = _load_audit_log()
    _claim_audit = [
        e for e in _all_audit
        if e.get("claim_id") == claim_id and e.get("sheet") == selected_sheet
    ]

    _ts_llm_unpack   = _dt.datetime.now()
    _llm_mappings    = (_llm_map_result or {}).get("mappings", {})
    _llm_reasoning   = (_llm_map_result or {}).get("_reasoning", {})
    _llm_called_at   = (_llm_map_result or {}).get("_timestamp", None)
    _llm_model       = (_llm_map_result or {}).get("_model", "see .env")
    _llm_reverse     = {v: k for k, v in _llm_mappings.items()}
    _llm_source_cols = set(_llm_mappings.keys())

    _ts_schema_map = _dt.datetime.now()
    _mapped: dict = {}
    if active_schema:
        from config.schemas import SCHEMAS
        if active_schema in SCHEMAS:
            _mapped = map_claim_to_schema(curr_claim, active_schema, {}, _llm_map_result)
    _ts_schema_map_done = _dt.datetime.now()
    _schema_map_ms = int((_ts_schema_map_done - _ts_schema_map).total_seconds() * 1000)

    # ── Pipeline trace ────────────────────────────────────────────────────────
    _pipeline_steps = []
    _step_row = (
        lambda color, icon, label, detail, ts:
        f"<div style='display:flex;align-items:center;gap:6px;padding:6px 0;"
        f"border-bottom:1px solid {_D_BDR2};'>"
        f"<span style='min-width:140px;font-size:10px;color:{color};font-weight:700;"
        f"font-family:monospace;'>{icon} {label}</span>"
        f"<span style='font-size:10px;color:{_D_LBL};font-family:monospace;'>→</span>"
        f"<span style='font-size:11px;color:{_D_TXT};font-family:monospace;flex:1;'>{detail}</span>"
        f"<span style='font-size:9px;color:{_D_LBL};font-family:monospace;white-space:nowrap;'>{ts}</span>"
        f"</div>"
    )

    _pipeline_steps.append(_step_row(
        _D_GRN, "📂", "FILE PARSED",
        "Claims read from the uploaded spreadsheet",
        _ts_fmt(_ts_dialog_open)
    ))
    if active_schema:
        _pipeline_steps.append(_step_row(
            _D_BLU, "🗂", "SCHEMA MAPPED",
            f"Fields matched to the {active_schema} schema template",
            f"{_ts_fmt(_ts_schema_map)} ({_schema_map_ms}ms)"
        ))
    if _llm_mappings:
        _llm_ts_display = _llm_called_at if _llm_called_at else _ts_fmt(_ts_llm_unpack)
        _pipeline_steps.append(_step_row(
            _D_YEL, "🤖", "LLM CALLED",
            f"AI resolved {len(_llm_mappings)} unrecognised column(s) to known fields",
            _llm_ts_display
        ))
    _edit_count = sum(1 for e in _claim_audit if e.get("event") == "FIELD_EDITED")
    if _edit_count:
        _last_edit_ts = max(
            (e.get("timestamp", "")[:19] for e in _claim_audit if e.get("event") == "FIELD_EDITED"),
            default=None,
        )
        _pipeline_steps.append(_step_row(
            _D_YEL, "✏", "USER EDITS",
            f"{_edit_count} field(s) manually updated by the user",
            (_last_edit_ts.replace("T", " ") if _last_edit_ts else "—")
        ))

    st.markdown(
        f"<div style='background:{_D_BG};border:1px solid {_D_BDR};border-left:4px solid {_D_YEL};"
        f"border-radius:8px;padding:12px 16px;margin-bottom:16px;"
        f"box-shadow:0 1px 4px rgba(0,0,0,0.06);'>"
        f"<div style='font-size:10px;font-weight:800;color:{_D_TXT};font-family:monospace;"
        f"text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;'>⚡ Pipeline Trace</div>"
        + "".join(_pipeline_steps)
        + "</div>",
        unsafe_allow_html=True,
    )

    # ── Field timeline header ─────────────────────────────────────────────────
    st.markdown(
        f"<div style='display:flex;align-items:center;gap:8px;margin-bottom:12px;'>"
        f"<div style='font-size:11px;font-weight:800;color:{_D_TXT};font-family:monospace;"
        f"text-transform:uppercase;letter-spacing:1.5px;'>Field Transformation Timeline</div>"
        f"<div style='flex:1;height:1px;background:linear-gradient(90deg,{_D_BDR},transparent);'></div>"
        f"</div>",
        unsafe_allow_html=True,
    )

    def _step_circle(n: int | str, color: str, bg: str) -> str:
        return (
            f"<div style='min-width:22px;height:22px;border-radius:50%;"
            f"background:{bg};border:2px solid {color};"
            f"display:flex;align-items:center;justify-content:center;"
            f"font-size:10px;color:{color};font-weight:bold;flex-shrink:0;'>{n}</div>"
        )

    fields_to_show = list(_mapped.keys()) if _mapped else list(curr_claim.keys())

    for field in fields_to_show:
        _ts_field = _dt.datetime.now()

        if _mapped and field in _mapped:
            m          = _mapped[field]
            raw_val    = m["info"].get("value", "")
            excel_col  = m.get("excel_field", field)
            hdr_score  = m.get("header_score", 0)
            val_score  = m.get("value_score", 0)
            conf       = m.get("confidence", 0)
            from_title = m.get("from_title", False)
            llm_mapped = bool(m.get("llm_mapped", False)) or (field in _llm_reverse)
            if llm_mapped and field in _llm_reverse:
                excel_col = _llm_reverse[field]
        else:
            if field not in curr_claim:
                continue
            info       = curr_claim[field]
            raw_val    = info.get("value", "")
            excel_col  = field
            hdr_score  = info.get("header_score", 0)
            val_score  = info.get("value_score", 0)
            conf       = info.get("confidence", 0)
            from_title = info.get("from_title", False)
            llm_mapped = (field in _llm_source_cols) or (field in _llm_reverse)

        mk_schema = f"mod_{selected_sheet}_{claim_id}_schema_{field}"
        mk_plain  = f"mod_{selected_sheet}_{claim_id}_{field}"
        cur_val   = st.session_state.get(mk_schema) or st.session_state.get(mk_plain) or raw_val
        edits     = _get_field_history(selected_sheet, claim_id, field)
        is_edited = cur_val != raw_val

        if from_title:
            method, method_color, method_icon = "TITLE ROW",   _D_PUR, "📋"
            method_fn = "parsing.py · extract_title_fields()"
        elif llm_mapped:
            method, method_color, method_icon = "LLM MAPPED",  _D_YEL, "🤖"
            method_fn = f"modules.llm · llm_map_unknown_fields() → {_llm_model}"
        elif hdr_score >= 90:
            method, method_color, method_icon = "EXACT MATCH", _D_GRN, "✓"
            method_fn = "modules.schema_mapping · _header_match_score()"
        elif hdr_score >= 65:
            method, method_color, method_icon = "FUZZY MATCH", _D_BLU, "~"
            method_fn = "modules.schema_mapping · _header_match_score() [fuzzy]"
        elif hdr_score > 0:
            method, method_color, method_icon = "PARTIAL MATCH", _D_LBL, "≈"
            method_fn = "modules.schema_mapping · _header_match_score() [partial]"
        else:
            method, method_color, method_icon = "DIRECT", _D_LBL, "→"
            method_fn = "modules.parsing · direct column read"

        conf_color   = _D_GRN if conf >= 80 else _D_YEL if conf >= 50 else _D_RED
        _display_val = raw_val if raw_val else f"<span style='color:{_D_LBL};'>(empty)</span>"
        _field_ts    = _ts_fmt(_ts_field)

        val_box_style = (
            f"font-size:13px;color:{_D_TXT};font-family:monospace;"
            f"background:{_D_BG2};border:1px solid {_D_BDR};border-radius:4px;"
            f"padding:4px 8px;margin-top:4px;word-break:break-all;"
        )
        code_style = f"background:{_D_BG2};padding:1px 5px;border-radius:3px;border:1px solid {_D_BDR};color:{_D_TXT};"

        steps_html = (
            # Step 1 — Extraction
            f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:8px;'>"
            f"{_step_circle(1, _D_GRN, _D_GRN_BG)}"
            f"<div style='flex:1;'>"
            f"<div style='display:flex;align-items:center;gap:8px;'>"
            f"<div style='font-size:11px;font-weight:700;color:{_D_GRN};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;'>Extracted from Document</div>"
            f"<span style='font-size:9px;color:{_D_LBL};font-family:monospace;margin-left:auto;'>"
            f"⏱ {_field_ts} · modules.parsing</span></div>"
            f"<div style='font-size:12px;color:{_D_LBL};margin-top:2px;'>"
            f"Column: <code style='{code_style}'>{excel_col}</code></div>"
            f"<div style='{val_box_style}'>{_display_val}</div>"
            f"</div></div>"
            f"<div style='margin-left:11px;border-left:2px dashed {_D_BDR};height:8px;margin-bottom:8px;'></div>"
            # Step 2 — Mapping
            f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:8px;'>"
            f"{_step_circle(2, method_color, _D_BLU_BG)}"
            f"<div style='flex:1;'>"
            f"<div style='display:flex;align-items:center;gap:8px;'>"
            f"<div style='font-size:11px;font-weight:700;color:{method_color};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;'>{method_icon} {method}</div>"
            f"<span style='font-size:9px;color:{_D_LBL};font-family:monospace;margin-left:auto;'>"
            f"⏱ {_field_ts} · {method_fn}</span></div>"
        )
        if llm_mapped:
            _src_col = _llm_reverse.get(field, field)
            _reason  = _llm_reasoning.get(_src_col, "")
            steps_html += (
                f"<div style='font-size:12px;color:{_D_LBL};margin-top:2px;'>"
                f"Source: <code style='{code_style}'>{_src_col}</code>"
                + (f" → <code style='{code_style}'>{field}</code>" if _src_col != field else "")
                + f"</div>"
            )
            if _reason:
                steps_html += (
                    f"<div style='font-size:11px;color:{_D_LBL};font-style:italic;"
                    f"margin-top:3px;padding:4px 8px;background:{_D_BG2};"
                    f"border-left:2px solid {_D_YEL};border-radius:0 4px 4px 0;'>"
                    f"LLM reasoning: {_reason}</div>"
                )
        elif hdr_score > 0:
            steps_html += (
                f"<div style='font-size:12px;color:{_D_LBL};margin-top:2px;'>"
                f"Header similarity: <span style='color:{method_color};font-weight:700;'>{hdr_score}%</span> · "
                f"Value quality: <span style='color:{conf_color};font-weight:700;'>{val_score}%</span> · "
                f"Overall confidence: <span style='color:{conf_color};font-weight:700;'>{conf}%</span></div>"
            )
        else:
            steps_html += (
                f"<div style='font-size:12px;color:{_D_LBL};margin-top:2px;'>"
                f"Column name matches field directly</div>"
            )
        steps_html += "</div></div>"

        for i, edit in enumerate(edits):
            steps_html += (
                f"<div style='margin-left:11px;border-left:2px dashed {_D_BDR};height:8px;margin-bottom:8px;'></div>"
                f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:8px;'>"
                f"{_step_circle(i+3, _D_YEL, _D_YEL_BG)}"
                f"<div style='flex:1;'>"
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<div style='font-size:11px;font-weight:700;color:{_D_YEL};font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1px;'>✏ User Edit</div>"
                f"<span style='font-size:9px;color:{_D_LBL};font-family:monospace;margin-left:auto;'>"
                f"⏱ {edit['ts']}</span></div>"
                f"<div style='display:flex;gap:8px;margin-top:4px;align-items:center;'>"
                f"<div style='font-size:12px;color:{_D_RED};font-family:monospace;"
                f"background:{_D_RED_BG};border:1px solid {_D_RED}60;border-radius:4px;"
                f"padding:3px 8px;word-break:break-all;flex:1;'>"
                f"<span style='font-size:10px;'>FROM: </span>{edit['from']}</div>"
                f"<div style='color:{_D_LBL};font-size:14px;'>→</div>"
                f"<div style='font-size:12px;color:{_D_GRN};font-family:monospace;"
                f"background:{_D_GRN_BG};border:1px solid {_D_GRN}60;border-radius:4px;"
                f"padding:3px 8px;word-break:break-all;flex:1;'>"
                f"<span style='font-size:10px;'>TO: </span>{edit['to']}</div>"
                f"</div></div></div>"
            )

        if is_edited:
            steps_html += (
                f"<div style='margin-left:11px;border-left:2px dashed {_D_BDR};height:8px;margin-bottom:8px;'></div>"
                f"<div style='display:flex;align-items:flex-start;gap:10px;margin-bottom:4px;'>"
                f"{_step_circle('✓', _D_GRN, _D_GRN_BG)}"
                f"<div style='flex:1;'>"
                f"<div style='font-size:11px;font-weight:700;color:{_D_GRN};font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1px;'>Final Value</div>"
                f"<div style='{val_box_style}color:{_D_GRN};border-color:{_D_GRN}60;'>{cur_val}</div>"
                f"</div></div>"
            )

        border_c = _D_YEL if is_edited else _D_BDR
        bg_c     = _D_YEL_BG if is_edited else _D_BG
        mod_badge = (
            f"<span style='font-size:9px;font-weight:700;color:{_D_YEL};"
            f"background:{_D_YEL_BG};border:1px solid {_D_YEL}60;"
            f"border-radius:10px;padding:2px 8px;font-family:monospace;"
            f"margin-left:8px;'>MODIFIED</span>" if is_edited else ""
        )
        st.markdown(
            f"<div style='background:{bg_c};border:1px solid {border_c};"
            f"border-radius:8px;padding:12px 14px;margin-bottom:10px;"
            f"box-shadow:0 1px 4px rgba(0,0,0,0.06);'>"
            f"<div style='font-size:12px;font-weight:800;color:{_D_TXT};font-family:monospace;"
            f"text-transform:uppercase;letter-spacing:1px;margin-bottom:10px;"
            f"display:flex;align-items:center;gap:8px;'>"
            f"{field}{mod_badge}</div>{steps_html}</div>",
            unsafe_allow_html=True,
        )

    # ── Audit log ─────────────────────────────────────────────────────────────
    if _claim_audit:
        st.markdown("---")

        _ev_cfg = {
            "FIELD_EDITED":             (_D_BLU,  "✏",  "Field edited"),
            "FIELD_ADDED":              (_D_PUR,  "＋", "Custom field added"),
            "EXPORT_GENERATED":         (_D_GRN,  "⬇",  "Export generated"),
            "FILE_UPLOADED":            (_D_YEL,  "📂", "File uploaded"),
            "SCHEMA_CHANGED":           (_D_LBL,  "🗂",  "Schema changed"),
            "LLM_CAUSE_ENRICHED":       (_D_LBL,  "🤖", "LLM enriched"),
            "CLAIM_DUPLICATE_DETECTED": (_D_RED,  "⚠",  "Duplicate detected"),
        }

        _USER_EVENTS  = {"FIELD_EDITED", "FIELD_ADDED", "EXPORT_GENERATED"}
        _session_start = st.session_state.get("_session_start", "")

        _seen_llm = False
        _deduped_audit: list = []
        for _e in _claim_audit:
            if _e.get("event") == "LLM_CAUSE_ENRICHED":
                if not _seen_llm:
                    _deduped_audit.append(_e)
                    _seen_llm = True
            else:
                _deduped_audit.append(_e)

        _session_user_events = [
            e for e in _deduped_audit
            if e.get("event") in _USER_EVENTS and e.get("timestamp", "") >= _session_start
        ]
        _full_events = _deduped_audit
        _show_full   = st.session_state[_full_hist_key]

        _type_counts: dict[str, int] = {}
        for _e in _session_user_events:
            _t = _e.get("event", "EVENT")
            _type_counts[_t] = _type_counts.get(_t, 0) + 1

        _summary_pills = "".join(
            f"<span style='background:{_ev_cfg.get(_t,(_D_LBL,'•',''))[0]}18;"
            f"border:1px solid {_ev_cfg.get(_t,(_D_LBL,'•',''))[0]}55;"
            f"border-radius:20px;padding:2px 8px;font-size:10px;"
            f"color:{_ev_cfg.get(_t,(_D_LBL,'•',_t))[0]};font-family:monospace;margin-right:4px;'>"
            f"{_ev_cfg.get(_t,(_D_LBL,'•',_t))[1]} {_n} {_t.replace('_',' ').lower()}</span>"
            for _t, _n in _type_counts.items()
        )

        _hdr_col, _btn_col = st.columns([7, 3])
        with _hdr_col:
            st.markdown(
                f"<div style='display:flex;align-items:center;gap:10px;margin-bottom:6px;'>"
                f"<div style='font-size:11px;font-weight:800;color:{_D_TXT};font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1.5px;'>📋 Audit Log"
                f"<span style='font-size:9px;color:{_D_LBL};margin-left:6px;font-weight:400;'>"
                f"(this session)</span></div>"
                f"<div style='flex:1;'>{_summary_pills}</div>"
                f"</div>",
                unsafe_allow_html=True,
            )
        with _btn_col:
            def _toggle_full_hist(_key=_full_hist_key):
                st.session_state[_key] = not st.session_state[_key]
            st.button(
                "▲ Hide Full History" if _show_full else "▼ View Full History",
                key=f"toggle_full_hist_{claim_id}",
                use_container_width=True,
                on_click=_toggle_full_hist,
            )

        def _render_audit_rows(events: list, id_prefix: str) -> None:
            for _ei, _event in enumerate(events):
                _ev_type  = _event.get("event", "EVENT")
                _ev_ts    = _event.get("timestamp", "")[:19].replace("T", " ")
                _ev_field = _event.get("field", "")
                _ev_from  = _event.get("original", "")
                _ev_to    = _event.get("new_value", "")
                _ev_recs  = _event.get("records", "")
                _ev_etype = _event.get("export_type", "")
                _cfg      = _ev_cfg.get(_ev_type, (_D_LBL, "•", _ev_type))
                _ev_color = _cfg[0]
                _ev_icon  = _cfg[1]

                _detail = ""
                if _ev_type == "FIELD_EDITED" and _ev_field:
                    _sf   = str(_ev_from)[:22] + ("…" if len(str(_ev_from)) > 22 else "")
                    _st_v = str(_ev_to)[:22]   + ("…" if len(str(_ev_to))   > 22 else "")
                    _detail = (
                        f"<span style='color:{_D_TXT};font-weight:600;'>{_ev_field}</span> "
                        f"<span style='color:{_D_RED};'>{_sf}</span>"
                        f"<span style='color:{_D_LBL};'> → </span>"
                        f"<span style='color:{_D_GRN};'>{_st_v}</span>"
                    )
                elif _ev_type == "FIELD_ADDED" and _ev_field:
                    _detail = f"<span style='color:{_D_PUR};font-weight:600;'>{_ev_field}</span>"
                elif _ev_type == "EXPORT_GENERATED":
                    _detail = (
                        f"<span style='color:{_D_GRN};font-weight:600;'>{_ev_etype}</span>"
                        + (f"<span style='color:{_D_LBL};'> · {_ev_recs} records</span>" if _ev_recs else "")
                    )
                elif _ev_type == "LLM_CAUSE_ENRICHED":
                    _cause = _event.get("cause_of_loss", "")
                    _detail = (
                        f"<span style='color:{_D_LBL};'>cause: </span>"
                        f"<span style='color:{_D_TXT};'>{_cause}</span>"
                        if _cause else f"<span style='color:{_D_LBL};font-style:italic;'>first enrichment only</span>"
                    )

                _card_key = f"{id_prefix}_{_ei}"
                _expanded = _card_key in st.session_state[_audit_expand_key]

                def _toggle_card(_ck=_card_key, _ek=_audit_expand_key):
                    if _ck in st.session_state[_ek]:
                        st.session_state[_ek].discard(_ck)
                    else:
                        st.session_state[_ek].add(_ck)

                _row_col, _xbtn_col = st.columns([10, 1])
                with _row_col:
                    st.markdown(
                        f"<div style='background:{_D_BG};border:1px solid {_D_BDR};"
                        f"border-left:3px solid {_ev_color};border-radius:6px;"
                        f"padding:8px 12px;font-family:monospace;font-size:11px;"
                        f"display:flex;align-items:center;gap:8px;'>"
                        f"<span style='color:{_ev_color};font-weight:700;min-width:16px;'>{_ev_icon}</span>"
                        f"<span style='color:{_ev_color};font-weight:700;min-width:175px;"
                        f"white-space:nowrap;'>{_ev_type}</span>"
                        f"<span style='color:{_D_LBL};'>·</span>"
                        f"<span style='color:{_D_LBL};min-width:135px;white-space:nowrap;'>{_ev_ts}</span>"
                        f"<span style='color:{_D_LBL};'>·</span>"
                        f"<span style='flex:1;'>{_detail}</span>"
                        f"</div>",
                        unsafe_allow_html=True,
                    )
                with _xbtn_col:
                    st.button(
                        "▲" if _expanded else "▼",
                        key=f"audit_expand_{_card_key}",
                        use_container_width=True,
                        help="Expand / collapse full details",
                        on_click=_toggle_card,
                    )

                if _expanded:
                    _detail_rows = ""
                    for _k, _v in _event.items():
                        if _k == "event" or _v in (None, "", []):
                            continue
                        _vs = str(_v)
                        if _k == "original":               _vc = _D_RED
                        elif _k == "new_value":            _vc = _D_GRN
                        elif _k == "timestamp":            _vc = _D_LBL
                        elif _k in ("field", "export_type"): _vc = _D_TXT
                        else:                              _vc = _D_LBL
                        _detail_rows += (
                            f"<div style='display:flex;gap:12px;padding:4px 0;"
                            f"border-bottom:1px solid {_D_BDR2};'>"
                            f"<span style='min-width:120px;font-size:10px;color:{_D_LBL};"
                            f"font-family:monospace;text-transform:uppercase;'>{_k}</span>"
                            f"<span style='font-size:11px;color:{_vc};font-family:monospace;"
                            f"word-break:break-all;'>{_vs}</span>"
                            f"</div>"
                        )
                    st.markdown(
                        f"<div style='background:{_D_BG2};border:1px solid {_ev_color}44;"
                        f"border-left:3px solid {_ev_color};border-radius:0 0 6px 6px;"
                        f"padding:10px 14px;margin-top:-6px;margin-bottom:8px;'>"
                        f"{_detail_rows}</div>",
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown("<div style='margin-bottom:6px;'></div>", unsafe_allow_html=True)

        if not _session_user_events:
            st.markdown(
                f"<div style='color:{_D_LBL};font-size:12px;font-family:monospace;"
                f"padding:10px 0;font-style:italic;'>"
                "No user actions recorded in this session yet for this claim.</div>",
                unsafe_allow_html=True,
            )
        else:
            _render_audit_rows(_session_user_events, f"user_{selected_sheet}_{claim_id}")

        if _show_full:
            _llm_count_raw = sum(1 for e in _claim_audit if e.get("event") == "LLM_CAUSE_ENRICHED")
            _suppressed    = _llm_count_raw - (1 if _llm_count_raw > 0 else 0)
            _note = f" · {_suppressed} duplicate LLM events suppressed" if _suppressed > 0 else ""
            st.markdown(
                f"<div style='background:{_D_BG};border:1px solid {_D_BDR};border-radius:8px;"
                f"padding:10px 14px;margin-top:8px;'>"
                f"<div style='font-size:10px;font-weight:800;color:{_D_TXT};font-family:monospace;"
                f"text-transform:uppercase;letter-spacing:1.5px;margin-bottom:10px;'>"
                f"🕓 Full History — {len(_full_events)} event(s){_note}</div>",
                unsafe_allow_html=True,
            )
            _render_audit_rows(_full_events, f"full_{selected_sheet}_{claim_id}")
            st.markdown("</div>", unsafe_allow_html=True)

    if st.button("Close", type="primary", use_container_width=True):
        st.session_state.pop("_open_journey_dialog", None)
        st.rerun()